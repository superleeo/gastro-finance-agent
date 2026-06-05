"""
kassenabrechnung.py — 德国餐厅月度 Kassenabrechnung 生成器
===========================================================
根据德国餐饮业标准格式，自动生成月度现金账报表 (Excel)。
模板来源于真实餐厅 (Japanisches Restaurant OSAKA)，已通用化。

输出:
  - Excel: kassenabrechnung{YYYYMM}.xlsx (精确复刻参考格式)
  - PDF:   通过 Excel → PDF 转换 (需 LibreOffice/Excel)

数据来源:
  - config.RestaurantConfig → 餐厅名称、地址
  - financial_analytics.DailyFinancials → 日度营业额
  - reconciliation → 支出记录
"""

from __future__ import annotations

import calendar
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side, numbers,
)
from openpyxl.utils import get_column_letter

from config import (
    DIR_OUTPUT,
    TAX_RATE_SPEISEN,
    TAX_RATE_GETRAENKE,
    restaurant as cfg,
    logger,
)
from financial_analytics import DailyFinancials, generate_demo_data


# ╔══════════════════════════════════════════════════════════╗
# ║                   Constants & Styling                    ║
# ╚══════════════════════════════════════════════════════════╝

# 列映射（参考模板）
COL_DATE      = "A"   # Datum
COL_BELEG     = "B"   # Beleg Nr.
COL_VORGANG   = "D"   # Vorgang (Einnahme / Ausgaben / 供应商名)
COL_EINN_19   = "F"   # Einnahme 19%
COL_EINN_7    = "G"   # Einnahme 7%
COL_AUSG_7    = "I"   # Ausgaben 7% (a.Haus)
COL_AUSG_BRUTTO = "J"  # Ausgaben Brutto
COL_AUSG_VST  = "K"   # Vorsteuer
COL_AUSG_NETTO = "L"  # Netto
COL_KASSE     = "M"   # Kassenstand (Gegenkonto)

# 样式
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FONT = Font(name="Calibri", size=11, bold=True)
TITLE_FONT  = Font(name="Calibri", size=14, bold=True)
BODY_FONT   = Font(name="Calibri", size=10)
SMALL_FONT  = Font(name="Calibri", size=9)
EURO_FORMAT = '#,##0.00'
DATE_FORMAT = 'YYYY-MM-DD'


@dataclass
class DayEntry:
    """单日 Kassenabrechnung 条目"""
    datum: date
    einnahme_19: Decimal = Decimal("0")     # Einnahme 19%
    einnahme_7: Decimal = Decimal("0")      # Einnahme 7%
    ausgabe_7: Decimal = Decimal("0")       # Ausgabe 7% (a.Haus)
    ec_visa: Decimal = Decimal("0")         # EC/Visa 收款
    trinkgeld_karte: Decimal = Decimal("0") # 刷卡小费
    ausgaben_list: List[Tuple[str, Decimal]] = field(default_factory=list)  # (描述, 金额)
    privateinlage: Decimal = Decimal("0")    # 私人存入
    privatentnahme: Decimal = Decimal("0")  # 私人取出


@dataclass
class MonthlyReport:
    """月度 Kassenabrechnung 完整数据"""
    year: int
    month: int
    days: List[DayEntry] = field(default_factory=list)
    vortrag: Decimal = Decimal("0")          # 上月结转
    expenses_from_bank: List[Tuple[str, Decimal]] = field(default_factory=list)

    @property
    def total_einnahme_19(self) -> Decimal:
        return sum((d.einnahme_19 for d in self.days), Decimal("0"))
    @property
    def total_einnahme_7(self) -> Decimal:
        return sum((d.einnahme_7 for d in self.days), Decimal("0"))
    @property
    def total_einnahme(self) -> Decimal:
        return self.total_einnahme_19 + self.total_einnahme_7
    @property
    def total_ausgaben(self) -> Decimal:
        total = sum((d.ausgabe_7 for d in self.days), Decimal("0"))
        for d in self.days:
            total += sum((amt for _, amt in d.ausgaben_list), Decimal("0"))
        return total
    @property
    def endbestand(self) -> Decimal:
        return self.vortrag + self.total_einnahme + sum(
            (d.ec_visa + d.privateinlage for d in self.days), Decimal("0")
        ) - self.total_ausgaben - sum(
            (d.privatentnahme for d in self.days), Decimal("0")
        )


# ╔══════════════════════════════════════════════════════════╗
# ║                   Excel Generator                        ║
# ╚══════════════════════════════════════════════════════════╝

def generate_kassenabrechnung(
    report: MonthlyReport,
    output_dir: Optional[Path] = None,
) -> Path:
    """
    生成标准格式的月度 Kassenabrechnung Excel 文件。

    参数
    ----
    report : MonthlyReport — 完整的月度数据
    output_dir : Path — 输出目录 (默认 DIR_OUTPUT)

    返回
    ----
    Path — 生成的 Excel 文件路径
    """
    if output_dir is None:
        output_dir = DIR_OUTPUT
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"kassenabrechnung{report.year}{report.month:02d}.xlsx"
    filepath = output_dir / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kassenabrechnung"

    # ── 列宽 ─────────────────────────────────────────────────
    col_widths = {"A": 22, "B": 10, "C": 2, "D": 28, "E": 2,
                  "F": 11, "G": 11, "H": 8, "I": 11, "J": 11,
                  "K": 11, "L": 11, "M": 13}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    row = 1  # 当前行号

    # ── 页眉 (rows 2-5) ──────────────────────────────────────
    row = 2
    ws[f"F{row}"] = cfg.name
    ws[f"F{row}"].font = Font(name="Calibri", size=12, bold=True)
    row = 3
    ws[f"F{row}"] = cfg.address if cfg.address else "R7,31"  # 税号占位
    ws[f"F{row}"].font = SMALL_FONT
    row = 4
    city = cfg.address.split(",")[-1].strip() if cfg.address and "," in cfg.address else "68161 Mannheim"
    ws[f"F{row}"] = city
    ws[f"F{row}"].font = SMALL_FONT

    # 标题行
    row = 5
    first_day = date(report.year, report.month, 1)
    last_day = date(report.year, report.month, calendar.monthrange(report.year, report.month)[1])
    ws[f"A{row}"] = "Kassenabrechnung"
    ws[f"A{row}"].font = TITLE_FONT
    ws[f"H{row}"] = f"Vom {first_day.strftime('%d.%m.%Y')} bis {last_day.strftime('%d.%m.%Y')}"
    ws[f"M{row}"] = "Seite 1"

    # ── 表头 (row 7-8) ───────────────────────────────────────
    row = 6  # 空行
    row = 7
    headers = {COL_DATE: "Datum", COL_BELEG: "Beleg Nr.",
               COL_VORGANG: "Vorgang",
               COL_EINN_19: "Einnahme", COL_EINN_7: "",
               COL_AUSG_7: "Ausgaben",
               COL_AUSG_BRUTTO: "", COL_AUSG_VST: "", COL_AUSG_NETTO: "",
               COL_KASSE: "Gegenkonto"}
    for col, h in headers.items():
        cell = ws[f"{col}{row}"]
        cell.value = h
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row = 8
    sub_headers = {
        COL_VORGANG: "Übertrag/",
        COL_EINN_19: "0.19", COL_EINN_7: "0.07",
        COL_AUSG_7: "0.07", COL_AUSG_BRUTTO: "Brutto",
        COL_AUSG_VST: "Vorsteuer", COL_AUSG_NETTO: "Netto",
        COL_KASSE: "Kassenstand",
    }
    for col in [COL_VORGANG, COL_EINN_19, COL_EINN_7, COL_AUSG_7,
                COL_AUSG_BRUTTO, COL_AUSG_VST, COL_AUSG_NETTO, COL_KASSE]:
        cell = ws[f"{col}{row}"]
        cell.value = sub_headers.get(col, "")
        cell.font = SMALL_FONT
        cell.border = THIN_BORDER

    # 表头行样式
    for r in [7, 8]:
        for c in [COL_DATE, COL_BELEG, COL_VORGANG, COL_EINN_19, COL_EINN_7,
                  COL_AUSG_7, COL_AUSG_BRUTTO, COL_AUSG_VST, COL_AUSG_NETTO, COL_KASSE]:
            ws[f"{c}{r}"].border = THIN_BORDER
            ws[f"{c}{r}"].alignment = Alignment(horizontal="center", wrap_text=True)

    # ── 上月结转 (row 9-10) ──────────────────────────────────
    row = 9
    ws[f"{COL_VORGANG}{row}"] = "Kassenbestand des Vortages"
    ws[f"{COL_EINN_19}{row}"] = "i.Haus"
    ws[f"{COL_EINN_7}{row}"] = "i.Haus"
    ws[f"{COL_AUSG_7}{row}"] = "a.Haus"
    for c in [COL_VORGANG, COL_EINN_19, COL_EINN_7, COL_AUSG_7]:
        ws[f"{c}{row}"].font = SMALL_FONT
    row = 10
    ws[f"{COL_VORGANG}{row}"] = float(report.vortrag)
    ws[f"{COL_VORGANG}{row}"].number_format = EURO_FORMAT
    ws[f"{COL_KASSE}{row}"] = float(report.vortrag)
    ws[f"{COL_KASSE}{row}"].number_format = EURO_FORMAT

    # ── 每日条目 ──────────────────────────────────────────────
    kassenstand = report.vortrag
    current_row = 11  # 当前写入行

    for day_entry in report.days:
        d = day_entry.datum
        r = current_row

        # 日期行
        ws[f"{COL_DATE}{r}"] = d
        ws[f"{COL_DATE}{r}"].number_format = DATE_FORMAT
        ws[f"{COL_VORGANG}{r}"] = "Einnahme"
        ws[f"{COL_EINN_19}{r}"] = float(day_entry.einnahme_19)
        ws[f"{COL_EINN_19}{r}"].number_format = EURO_FORMAT
        ws[f"{COL_EINN_7}{r}"] = float(day_entry.einnahme_7)
        ws[f"{COL_EINN_7}{r}"].number_format = EURO_FORMAT
        ws[f"{COL_AUSG_7}{r}"] = float(day_entry.ausgabe_7)
        ws[f"{COL_AUSG_7}{r}"].number_format = EURO_FORMAT

        # 样式
        for c in [COL_DATE, COL_VORGANG, COL_EINN_19, COL_EINN_7, COL_AUSG_7]:
            ws[f"{c}{r}"].font = BODY_FONT
            ws[f"{c}{r}"].border = THIN_BORDER
        ws[f"{COL_KASSE}{r}"].border = THIN_BORDER

        # 当日收入 = 19% + 7% + EC/Visa
        kassenstand += day_entry.einnahme_19 + day_entry.einnahme_7 + day_entry.ec_visa

        r += 1
        # 刷卡小费
        if day_entry.trinkgeld_karte > 0:
            ws[f"{COL_VORGANG}{r}"] = "Trinkgeld aus Kartenzahlung"
            ws[f"{COL_VORGANG}{r}"].font = BODY_FONT
            # 小费记在 H 列（参考模板）
            ws[f"H{r}"] = float(day_entry.trinkgeld_karte)
            ws[f"H{r}"].number_format = EURO_FORMAT

        r += 1
        # EC/Visa
        ws[f"{COL_VORGANG}{r}"] = "EC/Euro/Visa"
        ws[f"{COL_VORGANG}{r}"].font = BODY_FONT
        ws[f"{COL_AUSG_BRUTTO}{r}"] = float(day_entry.ec_visa)
        ws[f"{COL_AUSG_BRUTTO}{r}"].number_format = EURO_FORMAT

        # 每日余额
        for rr in range(current_row, r + 1):
            ws[f"{COL_KASSE}{rr}"] = float(kassenstand)
            ws[f"{COL_KASSE}{rr}"].number_format = EURO_FORMAT
            ws[f"{COL_KASSE}{rr}"].border = THIN_BORDER

        current_row = r + 1

    # ── 支出部分 ──────────────────────────────────────────────
    sum_einnahme_19 = Decimal("0")
    sum_einnahme_7 = Decimal("0")
    sum_ec_visa = Decimal("0")
    sum_ausgaben = Decimal("0")

    # 收集汇总数据
    for d_entry in report.days:
        sum_einnahme_19 += d_entry.einnahme_19
        sum_einnahme_7 += d_entry.einnahme_7
        sum_ec_visa += d_entry.ec_visa

    # 汇总行 (在每日条目之后)
    r = current_row
    ws[f"{COL_EINN_19}{r}"] = float(sum_einnahme_19)
    ws[f"{COL_EINN_19}{r}"].number_format = EURO_FORMAT
    ws[f"{COL_EINN_7}{r}"] = float(sum_einnahme_7)
    ws[f"{COL_EINN_7}{r}"].number_format = EURO_FORMAT
    ws[f"{COL_AUSG_BRUTTO}{r}"] = float(sum_ec_visa)
    ws[f"{COL_AUSG_BRUTTO}{r}"].number_format = EURO_FORMAT
    ws[f"{COL_KASSE}{r}"] = float(kassenstand)
    ws[f"{COL_KASSE}{r}"].number_format = EURO_FORMAT
    for c in [COL_EINN_19, COL_EINN_7, COL_AUSG_BRUTTO, COL_KASSE]:
        ws[f"{c}{r}"].font = BODY_FONT
        ws[f"{c}{r}"].border = THIN_BORDER

    # 如果在 demo 模式下，补充一些模拟工资支出
    if cfg.demo_mode:
        r += 2
        demo_expenses = [
            ("Lohn Mitarbeiter A", Decimal("850.00")),
            ("Lohn Mitarbeiter B", Decimal("720.00")),
            ("Lohn Küchenhilfe", Decimal("650.00")),
            ("Lohn Servicekraft", Decimal("580.00")),
        ]
        # 支出项标题
        r_start = r
        ws[f"{COL_VORGANG}{r}"] = "Ausgaben"
        ws[f"{COL_VORGANG}{r}"].font = Font(name="Calibri", size=10, bold=True)
        r += 1
        for desc, amt in demo_expenses:
            ws[f"{COL_VORGANG}{r}"] = desc
            ws[f"{COL_VORGANG}{r}"].font = BODY_FONT
            ws[f"{COL_AUSG_BRUTTO}{r}"] = float(amt)
            ws[f"{COL_AUSG_BRUTTO}{r}"].number_format = EURO_FORMAT
            sum_ausgaben += amt
            r += 1

        # 支出汇总
        total_exp = sum_ausgaben + sum_ec_visa  # 结合真实 EC/Visa 数据
        ws[f"{COL_AUSG_BRUTTO}{r}"] = float(total_exp)
        ws[f"{COL_AUSG_BRUTTO}{r}"].number_format = EURO_FORMAT
        ws[f"{COL_AUSG_BRUTTO}{r}"].font = BODY_FONT
        kassenstand -= sum_ausgaben
        ws[f"{COL_KASSE}{r}"] = float(kassenstand)
        ws[f"{COL_KASSE}{r}"].number_format = EURO_FORMAT
        ws[f"{COL_KASSE}{r}"].font = BODY_FONT

        r += 1
        ws[f"{COL_VORGANG}{r}"] = "Ausgaben"
        ws[f"{COL_VORGANG}{r}"].font = Font(name="Calibri", size=10, bold=True)
        # 公式式样
        formule = f"{sum_einnahme_19}+{sum_einnahme_7}+{sum_ec_visa}-{sum_ausgaben}=?"
        ws[f"H{r}"] = formule
        ws[f"H{r}"].font = SMALL_FONT
        r += 1
        ws[f"{COL_VORGANG}{r}"] = "Kassenbestand"
        ws[f"{COL_VORGANG}{r}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"{COL_KASSE}{r}"] = float(kassenstand)
        ws[f"{COL_KASSE}{r}"].number_format = EURO_FORMAT

    # ── 月度汇总表 (列表底部) ─────────────────────────────────
    summary_start = r + 3
    r = summary_start
    ws.merge_cells(f"A{r}:L{r}")
    ws[f"A{r}"] = "Monatsübersicht"
    ws[f"A{r}"].font = Font(name="Calibri", size=12, bold=True)
    r += 1

    # 汇总表头
    sum_headers = [
        ("A", "Datum"), ("C", "7% i.Haus"), ("D", "19% i.Haus"),
        ("E", "7% a.Haus"), ("F", "Kassenschnitt"), ("G", "Gutschein"),
        ("H", "Überweis."), ("I", "Kreditkarte"), ("L", "Trinkgeld"),
    ]
    for col, h in sum_headers:
        ws[f"{col}{r}"] = h
        ws[f"{col}{r}"].font = Font(name="Calibri", size=9, bold=True)
        ws[f"{col}{r}"].border = THIN_BORDER
    r += 1

    # 每日明细
    for day_entry in report.days:
        ws[f"A{r}"] = day_entry.datum
        ws[f"A{r}"].number_format = DATE_FORMAT
        ws[f"A{r}"].font = SMALL_FONT
        # 假设 demo 数据
        ws[f"C{r}"] = float(day_entry.einnahme_7) if cfg.demo_mode else 0
        ws[f"D{r}"] = float(day_entry.einnahme_19) if cfg.demo_mode else 0
        for c in ["A", "C", "D"]:
            ws[f"{c}{r}"].border = THIN_BORDER
            ws[f"{c}{r}"].font = SMALL_FONT
        r += 1

    # 汇总行
    r += 1
    ws[f"C{r}"] = float(report.total_einnahme_7)
    ws[f"D{r}"] = float(report.total_einnahme_19)
    for c in ["C", "D"]:
        ws[f"{c}{r}"].font = Font(name="Calibri", size=9, bold=True)
        ws[f"{c}{r}"].border = THIN_BORDER

    # 税务计算
    r += 2
    tax_lines = [
        ("7% i.Haus", report.total_einnahme_7,
         (report.total_einnahme_7 / (Decimal("1") + TAX_RATE_SPEISEN)).quantize(Decimal("0.01")) if report.total_einnahme_7 > 0 else Decimal("0")),
        ("19% i.Haus", report.total_einnahme_19,
         (report.total_einnahme_19 / (Decimal("1") + TAX_RATE_GETRAENKE)).quantize(Decimal("0.01")) if report.total_einnahme_19 > 0 else Decimal("0")),
    ]
    for label, brutto, netto in tax_lines:
        ws[f"C{r}"] = label
        ws[f"C{r}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"D{r}"] = float(brutto)
        ws[f"E{r}"] = "Netto"
        ws[f"F{r}"] = float(netto)
        for c in ["C", "D", "E", "F"]:
            ws[f"{c}{r}"].font = Font(name="Calibri", size=9)
            ws[f"{c}{r}"].border = THIN_BORDER
        r += 1

    r += 1
    total_netto = sum(
        (report.total_einnahme_7 / (Decimal("1") + TAX_RATE_SPEISEN) +
         report.total_einnahme_19 / (Decimal("1") + TAX_RATE_GETRAENKE)).quantize(Decimal("0.01"))
        if report.total_einnahme_7 + report.total_einnahme_19 > 0 else Decimal("0")
    ) if False else (
        (report.total_einnahme_7 / (Decimal("1") + TAX_RATE_SPEISEN)).quantize(Decimal("0.01")) +
        (report.total_einnahme_19 / (Decimal("1") + TAX_RATE_GETRAENKE)).quantize(Decimal("0.01"))
    ) if (report.total_einnahme_7 + report.total_einnahme_19) > 0 else Decimal("0")

    # 简化上述计算
    n7 = (report.total_einnahme_7 / (Decimal("1") + TAX_RATE_SPEISEN)).quantize(Decimal("0.01"))
    n19 = (report.total_einnahme_19 / (Decimal("1") + TAX_RATE_GETRAENKE)).quantize(Decimal("0.01"))
    total_netto = n7 + n19
    total_ust = report.total_einnahme_7 + report.total_einnahme_19 - total_netto

    ws[f"E{r}"] = "Ge.Netto"
    ws[f"F{r}"] = float(total_netto)
    ws[f"E{r}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"F{r}"].font = Font(name="Calibri", size=10, bold=True)
    r += 1
    ws[f"E{r}"] = "Ge.umST"
    ws[f"F{r}"] = float(total_ust)
    ws[f"E{r}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"F{r}"].font = Font(name="Calibri", size=10, bold=True)

    # ── 打印设置 ─────────────────────────────────────────────
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # ── 冻结窗格 ─────────────────────────────────────────────
    ws.freeze_panes = "A9"

    wb.save(filepath)
    logger.info(f"Kassenabrechnung 已生成: {filepath}")
    return filepath


# ╔══════════════════════════════════════════════════════════╗
# ║                   Builder from System Data               ║
# ╚══════════════════════════════════════════════════════════╝

def build_report_from_daily_data(
    daily_data: List[DailyFinancials],
    year: int,
    month: int,
    vortrag: float = 429.68,
) -> MonthlyReport:
    """
    从 financial_analytics 的日度数据构建 Kassenabrechnung。

    参数
    ----
    daily_data : 日度财务数据列表
    year, month : 目标年月
    vortrag : 上月结转余额
    """
    report = MonthlyReport(year=year, month=month, vortrag=Decimal(str(vortrag)))

    for dd in daily_data:
        if dd.datum.year != year or dd.datum.month != month:
            continue
        day = DayEntry(
            datum=dd.datum,
            einnahme_19=dd.brutto_19,
            einnahme_7=dd.brutto_7,
            ec_visa=Decimal("0"),
            privatinlage=Decimal("0"),
            privatentnahme=Decimal("0"),
        )
        report.days.append(day)

    report.days.sort(key=lambda x: x.datum)
    return report


# ╔══════════════════════════════════════════════════════════╗
# ║                   PDF Export (via Excel)                 ║
# ╚══════════════════════════════════════════════════════════╝

def export_to_pdf(xlsx_path: Path, output_dir: Optional[Path] = None) -> Optional[Path]:
    """
    将 Excel 转换为 PDF（需要 LibreOffice 或 Microsoft Excel）。
    优先级: LibreOffice (命令行) > 手动提示
    """
    import subprocess
    if output_dir is None:
        output_dir = DIR_OUTPUT

    pdf_path = xlsx_path.with_suffix(".pdf")

    # 尝试 LibreOffice
    try:
        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf",
             "--outdir", str(output_dir), str(xlsx_path)],
            capture_output=True, timeout=60,
        )
        if result.returncode == 0 and pdf_path.exists():
            logger.info(f"PDF 已生成: {pdf_path}")
            return pdf_path
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass

    logger.warning("未找到 LibreOffice，请手动将 Excel 导出为 PDF")
    return None


# ╔══════════════════════════════════════════════════════════╗
# ║                   Quick Test                             ║
# ╚══════════════════════════════════════════════════════════╝

if __name__ == "__main__":
    from datetime import date as dt_date

    # 使用 demo 数据生成一份测试报表
    demo_data = generate_demo_data(90)
    report = build_report_from_daily_data(demo_data, year=2026, month=2, vortrag=429.68)

    xlsx_path = generate_kassenabrechnung(report)
    print(f"✅ Excel 已生成: {xlsx_path}")
    print(f"   总营业额 19%: €{report.total_einnahme_19:,.2f}")
    print(f"   总营业额 7%:  €{report.total_einnahme_7:,.2f}")
    print(f"   合计:         €{report.total_einnahme:,.2f}")
    print(f"   期末余额:     €{report.endbestand:,.2f}")

    # 尝试 PDF
    pdf_path = export_to_pdf(xlsx_path)
    if pdf_path:
        print(f"✅ PDF 已生成:  {pdf_path}")
    else:
        print("ℹ️  安装 LibreOffice 后可自动生成 PDF: brew install libreoffice")
